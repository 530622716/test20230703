import logging
import os
import platform
import re
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.styles import PatternFill

logging.basicConfig(level="DEBUG")

"""
1. 确认并修改源文件的列名，需要和CountryConfig的key保持一致（key列的列名为keys）
2. 确认并修改实际使用的多语言路径
3. 确认并修改实际要对比的多语言
4. 确认并修改实际使用的源文案sheet页名称
"""

# 需确认：实际使用的多语言路径
app_path = r"D:\Anker_autotest\多语言\0614-logservice"

#  需确认：实际待验证的语言
verify_country = ["ENG"]

#  需确认：实际的源文案sheet页名称
st_name = "nebulalogservice"


# 不同国家对应的目录文件
CountryConfig = {
    "ENG": "values",
  #  "AR": "values-ar-rEG",
   # "DE": "values-de-rDE",
    #"ES": "values-es-rES",
   # "FR": "values-fr-rFR",
   # "IT": "values-it-rIT",
   # "JP": "values-ja",
  #  "KO": "values-ko-rKR",
   # "PT": "values-pt-rBR",
  #  "ZH": "values-zh-rCN",
   # "ZH-TC": "values-zh-rTW"
}


def parse_data(country, app_path):
    if platform.system() == "Windows":
        path = app_path + r"\%s\strings.xml" % CountryConfig.get(country)
    elif platform.system() == "Darwin":
        path = app_path + "/%s/strings.xml" % CountryConfig.get(country)
    tree = ET.parse(path)
    return tree.getroot()


def getXMLValue(content, name):
    for item in content:
        if item.attrib.get("name", "") == name:
            return item.text


def read_xlsx(xlsx_name, app_path, sheet_name=None, verify_country=None):
    if sheet_name is None or verify_country is None:
        raise Exception("输入待验证的sheet_name或语言列表！")
    key_col = None
    country_col = None
    fill_pass = PatternFill("solid", fgColor="00FF00")
    fill_fail = PatternFill("solid", fgColor="FF0000")
    fill_xfail = PatternFill("solid", fgColor="FFFF00")
    fill_space = PatternFill("solid", fgColor="888888")

    wb = openpyxl.load_workbook(xlsx_name, data_only=True)
    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb.worksheets[0]
    values = tuple(sheet.values)
    first_row = values[0]
    first_row_strip = [i.strip() for i in first_row if i is not None]
    if "result_" + sheet.title in wb.sheetnames:
        wb.remove(wb["result_" + sheet.title])
    sheet_result = wb.create_sheet("result_" + sheet.title, index=0)
    sheet_result_col = 2
    for country in verify_country:
        logging.info(f"正在处理{country}语言...")
        if country not in first_row_strip:
            logging.warning(f"待验证Excel中不存在{country}语言！")
            continue
        for i, value in enumerate(first_row):
            if value is None:
                continue
            if value.lower().strip() == '索引':
                key_col = i
                sheet_result.cell(1, 1).value = value
                continue
            elif value.lower().strip() == country.lower():
                country_col = i
                break
        content = parse_data(country, app_path)
        sheet_result.cell(1, sheet_result_col).value = sheet.cell(1, country_col + 1).value
        for i in range(len(values) - 1):
            keyName = sheet.cell(i + 2, key_col + 1).value
            sheet_result.cell(i + 2, 1).value = keyName
            if keyName:
                target_value = sheet.cell(i + 2, country_col + 1).value
                value = getXMLValue(content, keyName)
                if value is None:
                    sheet_result.cell(i + 2, sheet_result_col).value = target_value
                    sheet_result.cell(i + 2, sheet_result_col + 1).value = "Key Not Find"
                    sheet_result.cell(i + 2, sheet_result_col + 1).fill = fill_fail
                    continue
                sheet_result.cell(i + 2, sheet_result_col).value = target_value
                sheet_result.cell(i + 2, sheet_result_col + 1).value = value
                if target_value:
                    if isinstance(target_value, int):
                        target_value = str(target_value)
                    value = value.replace("%@", "%s")
                    target_value = target_value.replace(chr(0xA0), " ")
                    value = value.replace(chr(0xA0), " ")
                    # 判断条件
                    if target_value == value or target_value == re.sub(r'\\(?!n)', "", value) or \
                            re.sub(r'\\(?!n)', "", target_value) == value or target_value.replace("\\n", '\n') == value:
                        sheet_result.cell(i + 2, sheet_result_col + 1).fill = fill_pass
                    elif value.strip() == target_value.strip() or \
                            target_value.strip() == re.sub(r'\\(?!n)', "", value).strip() or \
                            re.sub(r'\\(?!n)', "", target_value).strip() == value.strip() or \
                            target_value.replace("\\n", '\n').strip() == value.strip():
                        sheet_result.cell(i + 2, sheet_result_col + 1).fill = fill_xfail
                    elif "".join(value.split()) == "".join(target_value.split()) and \
                            re.findall(r"\s", value) == re.findall(r"\s", target_value):
                        sheet_result.cell(i + 2, sheet_result_col + 1).fill = fill_xfail
                    else:
                        sheet_result.cell(i + 2, sheet_result_col + 1).fill = fill_fail
                else:
                    sheet_result.cell(i + 2, sheet_result_col).value = "未提供比对文案"
                    sheet_result.cell(i + 2, sheet_result_col).fill = fill_space
        sheet_result_col += 2
        logging.info(f"{country}语言处理完毕。")
    sheet_result.cell(sheet.max_row + 2, 1).value = "绿色表示对比一致"
    sheet_result.cell(sheet.max_row + 2, 1).fill = fill_pass
    sheet_result.cell(sheet.max_row + 3, 1).value = "黄色表示空字符差异，人工对一下前后空格"
    sheet_result.cell(sheet.max_row + 3, 1).fill = fill_xfail
    sheet_result.cell(sheet.max_row + 4, 1).value = "红色表示不一致"
    sheet_result.cell(sheet.max_row + 4, 1).fill = fill_fail
    sheet_result.cell(sheet.max_row + 5, 1).value = "灰色表示产品未提供比对文案"
    sheet_result.cell(sheet.max_row + 5, 1).fill = fill_space
    wb.save(xlsx_name)


if __name__ == "__main__":
    # app_path = input("输入解压后app的地址（解包的目标文件夹）：")
    if not app_path:
        app_path = os.getcwd()

    # 第一个参数为脚步同目录的Excel文件，sheet_name填写需要验证的工作表的名称
    read_xlsx(r"result_android.xlsx", app_path, sheet_name=st_name, verify_country=verify_country)
